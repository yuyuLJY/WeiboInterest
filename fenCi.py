# encoding: utf-8
import jieba
import re
import xlrd
import xlwt
from collections import defaultdict # 字典，用于词频统计
from openpyxl import load_workbook#写入exce

def drawPicture():
    print("")

# 返回词的频率和词
def countWord(sentence):
    '''
        jieba.add_word('猪猪包')
    jieba.add_word('热巴')
    jieba.add_word('迪丽热巴')
    jieba.add_word('掌心包')
    jieba.add_word("炒鸡")
    jieba.add_word('关晓彤')
    jieba.add_word('口袋魔法')
    jieba.add_word('苏菲')
    # jieba.add_word("烈儿")
    jieba.add_word("天猫旗舰店")
    jieba.add_word("大姨妈")
    jieba.add_word("烈儿")
    jieba.add_word("小猪猪")
    jieba.add_word("姨妈巾")
    jieba.add_word("囤货")
    jieba.add_word("双十一")
    jieba.add_word("双十二")
    jieba.add_word("挺好")
    jieba.add_word("挺划算")
    jieba.add_word("挺新")
    jieba.add_word("双11")
    jieba.add_word("双12")
    '''
    words_dict_path = "weibo/词库.txt"
    jieba.load_userdict(words_dict_path)
    word = jieba.lcut(sentence)
    # 检验分词效果
    for i in range(len(word)):
        print("word:"+ word[i])

    stopwords = []
    f = open("weibo/stopword_comment.txt", 'r', encoding='UTF-8')  # 返回一个文件对象
    line = f.readline()  # 调用文件的 readline()方法
    while line:
        line = line.replace('\n', '')
        stopwords.append(line)
        line = f.readline()
    f.close()
    stopwords.append(" ")
    stopwords.append('哒')
    wordfrequency = defaultdict(int)
    for w in word:
        if w not in stopwords:
            wordfrequency[w] += 1

    for w in wordfrequency:
        print(w+" "+str(wordfrequency[w]))
    return wordfrequency

def writeToExel(wordfrequency):
    '''
    path = "苏菲分词.xlsx"
    # path = "E:/Do it/ZhangXinBao"
    file = load_workbook(path)  # 打开excel
    nsheet = file.create_sheet('frequency', index=0)  # 新建表
    style = xlwt.XFStyle()
    nsheet.cell(1, 1, 'word',style)  # 写入表头
    nsheet.cell(1, 2, 'frequency',style)  # 写入表头

    wordfrequency_order = sorted(wordfrequency.items(), key=lambda x: x[1], reverse=True)  # 把字典按词频降序排列

    for n in range(2, len(wordfrequency_order) + 2):  # 把降序后的词频统计结果写入excel
        nsheet.cell(n, 1, wordfrequency_order[n - 2][0],style)
        nsheet.cell(n, 2, wordfrequency_order[n - 2][1],style)

    file.save(path)
    '''
    print("评论"+str(len(wordfrequency)))
    style = xlwt.XFStyle()
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet('#new')
    wordfrequency_order = sorted(wordfrequency.items(), key=lambda x: x[1], reverse=True)  # 把字典按词频降序排列
    for n in range(1, len(wordfrequency_order) + 2):  # 把降序后的词频统计结果写入excel
        worksheet.write(n, 1, wordfrequency_order[n - 2][0],style)
        worksheet.write(n, 2, wordfrequency_order[n - 2][1])

    '''
    i = 1
    for w in wordfrequency:
        worksheet.write(i, 0, w, style)  # Outputs 5
        worksheet.write(i, 1, wordfrequency[w])  # Outputs 2
        i = i+1
    '''
    workbook.save('test.xls')

def main():
    # 逐行读取文件，并把评论拼接成一个str
    str = " "
    '''
    f = open("苏菲淘宝评价clean.txt",'r', encoding='UTF-8')  # 返回一个文件对象
    line = f.readline()  # 调用文件的 readline()方法
    while line:
        line = line.replace('\n', '')
        print(line.split("   ")[1])
        str = str + line.split("   ")[1]
        line = f.readline()
    f.close()
    '''

    comment_list = ['1876693235','1978351291','619880505','1428990944','1743027543','735658915','1562697291','1601455762',
                    '1634059993','523790789','1783376715','2007283277']
    mytext = ''
    for i in range(len(comment_list)):
        exfile = xlrd.open_workbook("weibo/comment/"+comment_list[i]+".xls")
        sheet1 = exfile.sheet_by_name('#')  # 读取Sheet1的内容，根据实际情况填写表名

        n = sheet1.nrows  # 表的总行数
        for i in range(0, n):
            text = sheet1.row(i)[1].value  # 从第0行开始计数，第0行是栏目，第1行是要的内容
            mytext = mytext + " " + text  # 把每一天内容合并到一个str中

    print("组合后:"+mytext)
    mytext = re.sub('[a-zA-Z0-9]','',mytext)
    # fre,word = countWord(str)
    fre = countWord(mytext) # 返回词频字典
    writeToExel(fre)

main()